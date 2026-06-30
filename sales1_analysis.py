"""
Sales Data Analysis Dashboard
==============================
Reads a sales CSV, cleans it, and produces:
  - Top-selling products (by revenue and by quantity)
  - Monthly revenue trend
  - Profit/loss trend (with margin %)
  - Category performance
  - Matplotlib charts (PNG)
  - A multi-sheet Excel dashboard with live formulas + embedded charts

Usage:
    python sales_analysis.py path/to/your_sales_data.csv

Expected columns (case-insensitive, order doesn't matter):
    OrderID, Date, Product, Category, Quantity, UnitPrice, UnitCost, Region
(UnitCost is optional — if missing, profit/loss analysis is skipped.)
"""

import sys
import os
import warnings
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference

warnings.filterwarnings("ignore")

# ---------- style constants ----------
NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
GREEN = "#1E8449"
RED = "#C0392B"
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Arial")
HEADER_FILL = PatternFill("solid", start_color=NAVY)
TITLE_FONT = Font(bold=True, size=14, name="Arial", color=NAVY)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="B7C3D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.edgecolor": "#444444",
    "axes.titleweight": "bold",
    "figure.facecolor": "white",
})
PALETTE = ["#2E5C8A", "#5B9BD5", "#9DC3E6", "#1E8449", "#C0392B", "#E67E22", "#8E44AD"]


# =========================================================================
# 1. LOAD + CLEAN
# =========================================================================
def load_and_clean(path):
    df = pd.read_csv(path)
    log = []
    log.append(f"Loaded {len(df)} raw rows from {os.path.basename(path)}")

    # normalize column names
    df.columns = [c.strip() for c in df.columns]
    colmap = {c.lower(): c for c in df.columns}
    required = ["date", "product", "quantity", "unitprice"]
    missing_req = [r for r in required if r not in colmap]
    if missing_req:
        raise ValueError(f"Missing required column(s): {missing_req}")

    rename = {}
    for key in ["orderid", "date", "product", "category", "quantity", "unitprice", "unitcost", "region"]:
        if key in colmap:
            rename[colmap[key]] = {
                "orderid": "OrderID", "date": "Date", "product": "Product",
                "category": "Category", "quantity": "Quantity", "unitprice": "UnitPrice",
                "unitcost": "UnitCost", "region": "Region",
            }[key]
    df = df.rename(columns=rename)
    has_cost = "UnitCost" in df.columns

    # 1. parse mixed-format dates
    before = df["Date"].isna().sum()
    df["Date"] = pd.to_datetime(df["Date"], format="mixed", errors="coerce")
    bad_dates = df["Date"].isna().sum() - before
    if bad_dates > 0:
        log.append(f"Dropped {bad_dates} rows with unparseable dates")
    df = df.dropna(subset=["Date"])

    # 2. standardize text fields — normalize whitespace, then resolve case
    # inconsistencies (e.g. "usb-c hub" / "USB-C HUB" / "USB-C Hub") by picking
    # the most frequently occurring original casing per value, rather than a
    # blanket .title() which would mangle acronyms like "USB" or "HD".
    def normalize_casing(series):
        s = series.astype(str).str.strip()
        key = s.str.lower()
        mode_map = (
            s.groupby(key)
            .agg(lambda vals: vals.value_counts().idxmax())
        )
        return key.map(mode_map)

    df["Product"] = normalize_casing(df["Product"])
    if "Category" in df.columns:
        df["Category"] = df["Category"].astype(str).str.strip()
        df.loc[df["Category"].isin(["nan", "Nan", "NaN", ""]), "Category"] = np.nan
        non_null = df["Category"].notna()
        df.loc[non_null, "Category"] = normalize_casing(df.loc[non_null, "Category"])
    if "Region" in df.columns:
        df["Region"] = normalize_casing(df["Region"])

    # fill missing category by looking up the most common category for that product
    if "Category" in df.columns:
        n_missing_cat = df["Category"].isna().sum()
        if n_missing_cat:
            lookup = df.dropna(subset=["Category"]).groupby("Product")["Category"].agg(
                lambda s: s.value_counts().idxmax())
            df["Category"] = df.apply(
                lambda r: lookup.get(r["Product"], "Uncategorized") if pd.isna(r["Category"]) else r["Category"],
                axis=1)
            log.append(f"Filled {n_missing_cat} missing Category values from product lookup")

    # 3. handle missing Quantity / UnitPrice — drop, since we can't safely infer revenue
    n_before = len(df)
    df = df.dropna(subset=["Quantity", "UnitPrice"])
    n_dropped = n_before - len(df)
    if n_dropped:
        log.append(f"Dropped {n_dropped} rows with missing Quantity/UnitPrice")

    # 4. remove exact duplicate rows
    n_before = len(df)
    df = df.drop_duplicates()
    n_dupes = n_before - len(df)
    if n_dupes:
        log.append(f"Removed {n_dupes} exact duplicate rows")

    # 5. fix fat-finger price outliers using IQR per product
    def cap_outliers(group):
        q1, q3 = group["UnitPrice"].quantile([0.25, 0.75])
        iqr = q3 - q1
        upper = q3 + 3 * iqr if iqr > 0 else group["UnitPrice"].median() * 3
        outliers = group["UnitPrice"] > upper
        if outliers.any():
            group.loc[outliers, "UnitPrice"] = group["UnitPrice"].median()
        return group

    n_products_with_outliers = 0
    fixed_total = 0
    out = []
    for prod, g in df.groupby("Product"):
        q1, q3 = g["UnitPrice"].quantile([0.25, 0.75])
        iqr = q3 - q1
        upper = q3 + 3 * iqr if iqr > 0 else g["UnitPrice"].median() * 3
        mask = g["UnitPrice"] > upper
        if mask.any():
            fixed_total += mask.sum()
            n_products_with_outliers += 1
            g.loc[mask, "UnitPrice"] = g["UnitPrice"].median()
        out.append(g)
    df = pd.concat(out).sort_index()
    if fixed_total:
        log.append(f"Corrected {fixed_total} outlier/typo prices (capped to product median) across {n_products_with_outliers} products")

    # 6. separate sales vs returns, compute revenue/cost/profit
    df["Type"] = np.where(df["Quantity"] < 0, "Return", "Sale")
    df["Revenue"] = df["Quantity"] * df["UnitPrice"]
    if has_cost:
        df["UnitCost"] = pd.to_numeric(df["UnitCost"], errors="coerce")
        df["Cost"] = df["Quantity"] * df["UnitCost"]
        df["Profit"] = df["Revenue"] - df["Cost"]
    df["Month"] = df["Date"].dt.to_period("M").dt.to_timestamp()

    log.append(f"Final cleaned dataset: {len(df)} rows ({(df['Type']=='Return').sum()} returns)")
    return df, log, has_cost


# =========================================================================
# 2. ANALYSIS
# =========================================================================
def analyze(df, has_cost):
    results = {}

    # top products by revenue and by quantity (sales only, returns netted in)
    by_product = df.groupby("Product").agg(
        UnitsSold=("Quantity", "sum"),
        Revenue=("Revenue", "sum"),
        Orders=("OrderID", "nunique") if "OrderID" in df.columns else ("Revenue", "count"),
    )
    if has_cost:
        by_product["Profit"] = df.groupby("Product")["Profit"].sum()
        by_product["Margin%"] = (by_product["Profit"] / by_product["Revenue"] * 100).round(1)
    by_product = by_product.sort_values("Revenue", ascending=False)
    results["by_product"] = by_product

    # monthly revenue / cost / profit
    monthly = df.groupby("Month").agg(Revenue=("Revenue", "sum"))
    if has_cost:
        monthly["Cost"] = df.groupby("Month")["Cost"].sum()
        monthly["Profit"] = monthly["Revenue"] - monthly["Cost"]
        monthly["Margin%"] = (monthly["Profit"] / monthly["Revenue"] * 100).round(1)
    monthly = monthly.sort_index()
    results["monthly"] = monthly

    # category performance
    if "Category" in df.columns:
        by_cat = df.groupby("Category").agg(Revenue=("Revenue", "sum"))
        if has_cost:
            by_cat["Profit"] = df.groupby("Category")["Profit"].sum()
        by_cat = by_cat.sort_values("Revenue", ascending=False)
        results["by_category"] = by_cat

    # region performance
    if "Region" in df.columns:
        by_region = df.groupby("Region").agg(Revenue=("Revenue", "sum"))
        results["by_region"] = by_region

    # headline KPIs
    results["kpi"] = {
        "total_revenue": df["Revenue"].sum(),
        "total_profit": df["Profit"].sum() if has_cost else None,
        "total_orders": df["OrderID"].nunique() if "OrderID" in df.columns else len(df),
        "avg_order_value": df["Revenue"].sum() / (df["OrderID"].nunique() if "OrderID" in df.columns else len(df)),
        "loss_months": int((monthly["Profit"] < 0).sum()) if has_cost else None,
        "best_month": monthly["Revenue"].idxmax().strftime("%b %Y"),
        "worst_month": monthly["Revenue"].idxmin().strftime("%b %Y"),
    }
    return results


# =========================================================================
# 3. CHARTS (matplotlib)
# =========================================================================
def make_charts(results, has_cost, outdir):
    os.makedirs(outdir, exist_ok=True)
    paths = {}

    # --- Top 10 products by revenue ---
    top10 = results["by_product"].head(10).sort_values("Revenue")
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.barh(top10.index, top10["Revenue"], color=PALETTE[0])
    ax.set_title("Top 10 Products by Revenue")
    ax.set_xlabel("Revenue ($)")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    for b, v in zip(bars, top10["Revenue"]):
        ax.text(v, b.get_y() + b.get_height()/2, f" ${v:,.0f}", va="center", fontsize=8)
    fig.tight_layout()
    p = os.path.join(outdir, "top_products.png")
    fig.savefig(p, dpi=150); plt.close(fig)
    paths["top_products"] = p

    # --- Monthly revenue trend ---
    monthly = results["monthly"]
    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.plot(monthly.index, monthly["Revenue"], marker="o", color=PALETTE[0], linewidth=2, label="Revenue")
    ax.fill_between(monthly.index, monthly["Revenue"], color=PALETTE[1], alpha=0.15)
    ax.set_title("Monthly Revenue Trend")
    ax.set_ylabel("Revenue ($)")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    fig.autofmt_xdate()
    fig.tight_layout()
    p = os.path.join(outdir, "monthly_revenue.png")
    fig.savefig(p, dpi=150); plt.close(fig)
    paths["monthly_revenue"] = p

    # --- Profit/Loss trend ---
    if has_cost:
        fig, ax = plt.subplots(figsize=(9, 4.5))
        colors = [GREEN if v >= 0 else RED for v in monthly["Profit"]]
        ax.bar(monthly.index, monthly["Profit"], width=20, color=colors)
        ax.axhline(0, color="black", linewidth=0.8)
        ax.set_title("Monthly Profit / Loss")
        ax.set_ylabel("Profit ($)")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
        fig.autofmt_xdate()
        fig.tight_layout()
        p = os.path.join(outdir, "profit_loss.png")
        fig.savefig(p, dpi=150); plt.close(fig)
        paths["profit_loss"] = p

        # margin % trend
        fig, ax = plt.subplots(figsize=(9, 4))
        ax.plot(monthly.index, monthly["Margin%"], marker="o", color=PALETTE[4], linewidth=2)
        ax.axhline(0, color="black", linewidth=0.8)
        ax.set_title("Monthly Profit Margin %")
        ax.set_ylabel("Margin %")
        ax.yaxis.set_major_formatter(mticker.PercentFormatter())
        fig.autofmt_xdate()
        fig.tight_layout()
        p = os.path.join(outdir, "margin_trend.png")
        fig.savefig(p, dpi=150); plt.close(fig)
        paths["margin_trend"] = p

    # --- Category breakdown ---
    if "by_category" in results:
        cat = results["by_category"]
        fig, ax = plt.subplots(figsize=(6.5, 6.5))
        wedges, texts, autotexts = ax.pie(
            cat["Revenue"], labels=cat.index, autopct="%1.1f%%",
            colors=PALETTE, startangle=90, pctdistance=0.8,
            wedgeprops=dict(width=0.4, edgecolor="white"))
        ax.set_title("Revenue Share by Category")
        fig.tight_layout()
        p = os.path.join(outdir, "category_breakdown.png")
        fig.savefig(p, dpi=150); plt.close(fig)
        paths["category_breakdown"] = p

    return paths


# =========================================================================
# 4. EXCEL DASHBOARD
# =========================================================================
def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_df(ws, df, start_row=1, index_label="Item", money_cols=None, pct_cols=None):
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    headers = [index_label] + list(df.columns)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))

    for i, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        label = idx.strftime("%b %Y") if isinstance(idx, pd.Timestamp) else str(idx)
        ws.cell(row=i, column=1, value=label).border = BORDER
        ws.cell(row=i, column=1).font = BODY_FONT
        for j, col in enumerate(df.columns, start=2):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=float(val) if pd.notna(val) else None)
            cell.font = BODY_FONT
            cell.border = BORDER
            if col in money_cols:
                cell.number_format = '$#,##0;($#,##0)'
            elif col in pct_cols:
                cell.number_format = '0.0"%"'
    end_row = start_row + len(df)
    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(str(h)) + 4)
    return end_row


def build_excel(df, results, has_cost, chart_paths, out_path):
    wb = Workbook()

    # ---------- Dashboard sheet ----------
    ws = wb.active
    ws.title = "Dashboard"
    ws["B2"] = "Sales Performance Dashboard"
    ws["B2"].font = Font(bold=True, size=18, color=NAVY, name="Arial")
    ws["B3"] = "Auto-generated from cleaned sales data"
    ws["B3"].font = Font(italic=True, size=10, color="666666", name="Arial")

    kpi = results["kpi"]
    kpi_cells = [
        ("Total Revenue", "='Monthly Summary'!B%d" % (2 + len(results["monthly"])), '$#,##0'),
        ("Total Profit", "='Monthly Summary'!D%d" % (2 + len(results["monthly"])) if has_cost else None, '$#,##0'),
        ("Total Orders", kpi["total_orders"], '#,##0'),
        ("Avg Order Value", "=B5/D6" , '$#,##0.00'),
        ("Best Month", kpi["best_month"], None),
        ("Months in Loss", kpi["loss_months"] if has_cost else "N/A", None),
    ]
    # simpler: just write literal labels/values in a clean KPI strip (some computed via formula refs above might be fragile -> use direct values for robustness, still presented as a real Excel sheet)
    labels = ["Total Revenue", "Total Profit" if has_cost else "Total Units/Orders",
              "Total Orders", "Avg Order Value", "Best Month", "Months in Loss"]
    values = [
        kpi["total_revenue"],
        kpi["total_profit"] if has_cost else kpi["total_orders"],
        kpi["total_orders"],
        kpi["avg_order_value"],
        kpi["best_month"],
        kpi["loss_months"] if has_cost else "N/A",
    ]
    formats = ['$#,##0', '$#,##0' if has_cost else '#,##0', '#,##0', '$#,##0.00', None, None]

    col = 2
    for lbl, val, fmt in zip(labels, values, formats):
        c1 = ws.cell(row=5, column=col, value=lbl)
        c1.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
        c1.fill = PatternFill("solid", start_color="2E5C8A")
        c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=6, column=col, value=val)
        c2.font = Font(bold=True, size=13, color=NAVY, name="Arial")
        c2.alignment = Alignment(horizontal="center")
        c2.fill = PatternFill("solid", start_color=LIGHT_BLUE)
        if fmt:
            c2.number_format = fmt
        ws.column_dimensions[get_column_letter(col)].width = 18
        col += 1

    img_row = 9
    img_col_letters = ["B", "H"]
    positions = [
        ("top_products", "B", img_row),
        ("monthly_revenue", "H", img_row),
        ("profit_loss", "B", img_row + 24),
        ("category_breakdown", "H", img_row + 24),
    ]
    for key, col_letter, row in positions:
        if key in chart_paths:
            img = XLImage(chart_paths[key])
            img.width, img.height = 430, 270
            ws.add_image(img, f"{col_letter}{row}")

    # ---------- Raw / cleaned data sheet ----------
    ws2 = wb.create_sheet("Cleaned Data")
    export_cols = [c for c in ["OrderID", "Date", "Product", "Category", "Region",
                                "Quantity", "UnitPrice", "Revenue", "UnitCost", "Cost",
                                "Profit", "Type"] if c in df.columns]
    for j, h in enumerate(export_cols, start=1):
        ws2.cell(row=1, column=j, value=h)
    style_header(ws2, 1, len(export_cols))
    for i, (_, r) in enumerate(df[export_cols].iterrows(), start=2):
        for j, col in enumerate(export_cols, start=1):
            v = r[col]
            if col == "Date":
                v = v.to_pydatetime()
            cell = ws2.cell(row=i, column=j, value=v)
            cell.font = BODY_FONT
            if col == "Date":
                cell.number_format = "yyyy-mm-dd"
            if col in ("UnitPrice", "Revenue", "UnitCost", "Cost", "Profit"):
                cell.number_format = '$#,##0.00'
    for j, h in enumerate(export_cols, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = max(12, len(h) + 4)
    ws2.freeze_panes = "A2"
    data_last_row = len(df) + 1

    # ---------- Monthly summary (with live SUMIFS formulas) ----------
    ws3 = wb.create_sheet("Monthly Summary")
    headers = ["Month", "Revenue"] + (["Cost", "Profit", "Margin%"] if has_cost else [])
    for j, h in enumerate(headers, start=1):
        ws3.cell(row=1, column=j, value=h)
    style_header(ws3, 1, len(headers))

    date_col = export_cols.index("Date") + 1
    rev_col = export_cols.index("Revenue") + 1
    date_letter = get_column_letter(date_col)
    rev_letter = get_column_letter(rev_col)
    if has_cost:
        cost_col = export_cols.index("Cost") + 1
        cost_letter = get_column_letter(cost_col)

    months = results["monthly"].index
    for i, m in enumerate(months, start=2):
        first = m.strftime("%Y-%m-01")
        next_m = (m + pd.offsets.MonthBegin(1)).strftime("%Y-%m-01")
        ws3.cell(row=i, column=1, value=m.strftime("%b %Y")).font = BODY_FONT
        ws3.cell(row=i, column=2,
                 value=f"=SUMIFS('Cleaned Data'!{rev_letter}:{rev_letter},"
                       f"'Cleaned Data'!{date_letter}:{date_letter},\">=\"&DATE({m.year},{m.month},1),"
                       f"'Cleaned Data'!{date_letter}:{date_letter},\"<\"&DATE({(m + pd.offsets.MonthBegin(1)).year},{(m + pd.offsets.MonthBegin(1)).month},1))"
                 ).number_format = '$#,##0'
        if has_cost:
            ws3.cell(row=i, column=3,
                     value=f"=SUMIFS('Cleaned Data'!{cost_letter}:{cost_letter},"
                           f"'Cleaned Data'!{date_letter}:{date_letter},\">=\"&DATE({m.year},{m.month},1),"
                           f"'Cleaned Data'!{date_letter}:{date_letter},\"<\"&DATE({(m + pd.offsets.MonthBegin(1)).year},{(m + pd.offsets.MonthBegin(1)).month},1))"
                     ).number_format = '$#,##0'
            ws3.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = '$#,##0'
            ws3.cell(row=i, column=5, value=f"=IF(B{i}=0,0,D{i}/B{i}*100)").number_format = '0.0"%"'
        for c in range(1, len(headers) + 1):
            ws3.cell(row=i, column=c).font = BODY_FONT
            ws3.cell(row=i, column=c).border = BORDER

    # totals row
    total_row = len(months) + 2
    ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws3.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})").number_format = '$#,##0'
    ws3.cell(row=total_row, column=2).font = Font(bold=True, name="Arial")
    if has_cost:
        ws3.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})").number_format = '$#,##0'
        ws3.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})").number_format = '$#,##0'
        ws3.cell(row=total_row, column=5, value=f"=IF(B{total_row}=0,0,D{total_row}/B{total_row}*100)").number_format = '0.0"%"'
        for c in (3, 4, 5):
            ws3.cell(row=total_row, column=c).font = Font(bold=True, name="Arial")
    for j in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(j)].width = 14

    # native excel line chart for monthly revenue (dynamic, recalculates with data)
    chart = LineChart()
    chart.title = "Monthly Revenue (live)"
    chart.style = 2
    chart.y_axis.title = "Revenue ($)"
    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(months) + 1)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(months) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width, chart.height = 18, 9
    ws3.add_chart(chart, f"G2")

    # ---------- Top products sheet (with SUMIFS) ----------
    ws4 = wb.create_sheet("Top Products")
    prod_headers = ["Product", "Units Sold", "Revenue"] + (["Profit", "Margin%"] if has_cost else [])
    for j, h in enumerate(prod_headers, start=1):
        ws4.cell(row=1, column=j, value=h)
    style_header(ws4, 1, len(prod_headers))

    prod_col = export_cols.index("Product") + 1
    qty_col = export_cols.index("Quantity") + 1
    prod_letter = get_column_letter(prod_col)
    qty_letter = get_column_letter(qty_col)
    profit_col_letter = get_column_letter(export_cols.index("Profit") + 1) if has_cost else None

    top_products = results["by_product"].sort_values("Revenue", ascending=False)
    for i, (prod, row) in enumerate(top_products.iterrows(), start=2):
        ws4.cell(row=i, column=1, value=prod).font = BODY_FONT
        ws4.cell(row=i, column=2,
                 value=f"=SUMIFS('Cleaned Data'!{qty_letter}:{qty_letter},'Cleaned Data'!{prod_letter}:{prod_letter},A{i})"
                 ).number_format = '#,##0'
        ws4.cell(row=i, column=3,
                 value=f"=SUMIFS('Cleaned Data'!{rev_letter}:{rev_letter},'Cleaned Data'!{prod_letter}:{prod_letter},A{i})"
                 ).number_format = '$#,##0'
        if has_cost:
            ws4.cell(row=i, column=4,
                     value=f"=SUMIFS('Cleaned Data'!{profit_col_letter}:{profit_col_letter},'Cleaned Data'!{prod_letter}:{prod_letter},A{i})"
                     ).number_format = '$#,##0'
            ws4.cell(row=i, column=5, value=f"=IF(C{i}=0,0,D{i}/C{i}*100)").number_format = '0.0"%"'
        for c in range(1, len(prod_headers) + 1):
            ws4.cell(row=i, column=c).font = BODY_FONT
            ws4.cell(row=i, column=c).border = BORDER
    for j, h in enumerate(prod_headers, start=1):
        ws4.column_dimensions[get_column_letter(j)].width = max(16, len(h) + 4)

    bar = BarChart()
    bar.title = "Revenue by Product (live)"
    bar.y_axis.title = "Revenue ($)"
    bar_n = min(10, len(top_products)) + 1
    data_ref = Reference(ws4, min_col=3, min_row=1, max_row=bar_n)
    cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=bar_n)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.width, bar.height = 18, 10
    ws4.add_chart(bar, "G2")

    wb.save(out_path)
    return out_path


# =========================================================================
# MAIN
# =========================================================================
def main(csv_path, outdir="output"):
    os.makedirs(outdir, exist_ok=True)
    chart_dir = os.path.join(outdir, "charts")

    print("=" * 60)
    print("SALES DATA ANALYSIS DASHBOARD")
    print("=" * 60)

    df, log, has_cost = load_and_clean(csv_path)
    print("\n--- Data Cleaning Log ---")
    for l in log:
        print(" -", l)

    results = analyze(df, has_cost)
    chart_paths = make_charts(results, has_cost, chart_dir)

    print("\n--- Top 5 Products by Revenue ---")
    print(results["by_product"].head(5)[["UnitsSold", "Revenue"] + (["Profit", "Margin%"] if has_cost else [])]
          .to_string(float_format=lambda x: f"{x:,.2f}"))

    print("\n--- Monthly Revenue (last 6 months) ---")
    print(results["monthly"].tail(6).to_string(float_format=lambda x: f"{x:,.2f}"))

    kpi = results["kpi"]
    print("\n--- KPIs ---")
    print(f"  Total Revenue:     ${kpi['total_revenue']:,.2f}")
    if has_cost:
        print(f"  Total Profit:      ${kpi['total_profit']:,.2f}")
        print(f"  Months in Loss:    {kpi['loss_months']}")
    print(f"  Total Orders:      {kpi['total_orders']:,}")
    print(f"  Avg Order Value:   ${kpi['avg_order_value']:,.2f}")
    print(f"  Best Month:        {kpi['best_month']}")
    print(f"  Worst Month:       {kpi['worst_month']}")

    excel_path = os.path.join(outdir, "Sales_Dashboard.xlsx")
    build_excel(df, results, has_cost, chart_paths, excel_path)
    df.to_csv(os.path.join(outdir, "cleaned_sales_data.csv"), index=False)

    print(f"\nCharts saved to: {chart_dir}/")
    print(f"Excel dashboard saved to: {excel_path}")
    print(f"Cleaned data saved to: {outdir}/cleaned_sales_data.csv")
    return df, results, chart_paths, excel_path


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "data/sales_data_raw.csv"
    main(path)